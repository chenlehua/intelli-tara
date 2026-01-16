"""Excel document parser using openpyxl."""

from typing import Any, List

from openpyxl import load_workbook

from app.services.parsers.base import BaseParser, ParsedContent


class ExcelParser(BaseParser):
    """Parser for Excel documents."""

    SUPPORTED_TYPES = ['xlsx', 'xls']

    def supports(self, file_type: str) -> bool:
        """Check if this parser supports the given file type."""
        return file_type.lower() in self.SUPPORTED_TYPES

    async def parse(self, file_path: str) -> ParsedContent:
        """Parse an Excel document.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            ParsedContent with extracted tables
        """
        result = ParsedContent()

        try:
            wb = load_workbook(file_path, data_only=True)

            result.metadata = {
                "sheet_names": wb.sheetnames,
                "sheet_count": len(wb.sheetnames),
            }

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Extract sheet data as table
                table_data = []
                for row in ws.iter_rows(values_only=True):
                    row_data = [self._cell_to_str(cell) for cell in row]
                    # Skip completely empty rows
                    if any(cell for cell in row_data):
                        table_data.append(row_data)

                if table_data:
                    result.tables.append(table_data)

                    # Also add as text block for searchability
                    sheet_text = f"Sheet: {sheet_name}\n"
                    for row in table_data[:10]:  # First 10 rows as preview
                        sheet_text += " | ".join(row) + "\n"
                    result.text_blocks.append(sheet_text)

            wb.close()

        except Exception as e:
            result.metadata["error"] = str(e)

        return result

    def _cell_to_str(self, cell: Any) -> str:
        """Convert a cell value to string."""
        if cell is None:
            return ""
        return str(cell).strip()
