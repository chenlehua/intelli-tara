"""TARA Report Generator Service.

Generates Excel reports following the MY25EV_Platform_IVI_TARA_Report.xlsx format.
"""

import os
from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from app.models.asset import Asset
from app.models.project import Project, ProjectConfig
from app.models.threat import SecurityMitigation, ThreatScenario


class TARAReportGenerator:
    """Generator for TARA analysis reports in Excel format."""

    def __init__(self):
        self.wb = Workbook()
        self._setup_styles()

    def _setup_styles(self):
        """Setup common styles for the report."""
        self.header_font = Font(bold=True, size=11, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.subheader_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Risk level colors
        self.risk_colors = {
            1: PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),  # Green - Acceptable
            2: PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),  # Yellow - Low
            3: PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),  # Orange - Medium
            4: PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid"),  # Dark Orange - High
            5: PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),  # Red - Critical
        }

    async def generate(
        self,
        project: Project,
        config: Optional[ProjectConfig],
        assets: List[Asset],
        threats: List[ThreatScenario],
        output_dir: str,
    ) -> str:
        """Generate a complete TARA report.
        
        Args:
            project: Project model
            config: Project configuration
            assets: List of assets
            threats: List of threat scenarios
            output_dir: Directory to save the report
            
        Returns:
            Path to the generated report file
        """
        # Create sheets
        self._create_cover_sheet(project, config)
        self._create_definition_sheet(project, config)
        self._create_asset_sheet(assets)
        self._create_attack_tree_sheet(project)
        self._create_tara_result_sheet(threats, assets)

        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        # Save file
        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"TARA_Report_{project.id}_{timestamp}.xlsx"
        file_path = os.path.join(output_dir, filename)
        self.wb.save(file_path)

        return file_path

    def _create_cover_sheet(self, project: Project, config: Optional[ProjectConfig]):
        """Create the cover sheet."""
        ws = self.wb.create_sheet("封面", 0)

        # Title
        ws.merge_cells('A1:H3')
        title_cell = ws['A1']
        title_cell.value = "威胁分析与风险评估报告\nThreat Analysis and Risk Assessment Report"
        title_cell.font = Font(bold=True, size=24)
        title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Project info
        info_data = [
            ("项目名称", project.name),
            ("项目编码", project.code or ""),
            ("文档版本", "1.0"),
            ("编制日期", datetime.now().strftime("%Y-%m-%d")),
            ("编制", ""),
            ("审核", ""),
            ("会签", ""),
            ("批准", ""),
        ]

        for idx, (label, value) in enumerate(info_data, start=6):
            ws.cell(row=idx, column=2, value=label).font = Font(bold=True)
            ws.cell(row=idx, column=3, value=value)

        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40

    def _create_definition_sheet(self, project: Project, config: Optional[ProjectConfig]):
        """Create the definitions sheet."""
        ws = self.wb.create_sheet("相关定义", 1)

        row = 1

        # Section 1: Functional Description
        ws.cell(row=row, column=1, value="1. 功能描述").font = Font(bold=True, size=14)
        row += 1
        ws.cell(row=row, column=1, value=config.functional_description if config else "")
        ws.merge_cells(f'A{row}:H{row+3}')
        row += 5

        # Section 2: Project Boundary
        ws.cell(row=row, column=1, value="2. 项目边界").font = Font(bold=True, size=14)
        row += 1
        ws.cell(row=row, column=1, value=config.item_boundary if config else "")
        ws.merge_cells(f'A{row}:H{row+3}')
        row += 5

        # Section 3: System Architecture
        ws.cell(row=row, column=1, value="3. 系统架构图").font = Font(bold=True, size=14)
        row += 1
        ws.cell(row=row, column=1, value="[系统架构图占位]")
        row += 10

        # Section 4: Assumptions
        ws.cell(row=row, column=1, value="4. 相关假设").font = Font(bold=True, size=14)
        row += 1

        # Assumptions table header
        headers = ["序号", "假设内容", "说明"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
        row += 1

        # Section 5: Terminology
        row += 3
        ws.cell(row=row, column=1, value="5. 术语表").font = Font(bold=True, size=14)
        row += 1

        terms = [
            ("TARA", "Threat Analysis and Risk Assessment", "威胁分析与风险评估"),
            ("STRIDE", "Spoofing, Tampering, Repudiation, Information Disclosure, DoS, EoP", "威胁建模方法"),
            ("WP29", "World Forum for Harmonization of Vehicle Regulations", "联合国车辆法规协调论坛"),
        ]

        headers = ["术语", "英文全称", "说明"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
        row += 1

        for term, full_name, desc in terms:
            ws.cell(row=row, column=1, value=term).border = self.thin_border
            ws.cell(row=row, column=2, value=full_name).border = self.thin_border
            ws.cell(row=row, column=3, value=desc).border = self.thin_border
            row += 1

        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 30

    def _create_asset_sheet(self, assets: List[Asset]):
        """Create the asset list sheet."""
        ws = self.wb.create_sheet("资产列表", 2)

        # Headers
        headers = [
            "资产ID", "资产名称", "分类", "细分类", "备注",
            "真实性", "完整性", "不可抵赖性", "机密性", "可用性", "权限"
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.thin_border
            cell.alignment = self.center_align

        # Data rows
        for row_idx, asset in enumerate(assets, 2):
            ws.cell(row=row_idx, column=1, value=asset.asset_id).border = self.thin_border
            ws.cell(row=row_idx, column=2, value=asset.name).border = self.thin_border
            ws.cell(row=row_idx, column=3, value=asset.category).border = self.thin_border
            ws.cell(row=row_idx, column=4, value=asset.subcategory or "").border = self.thin_border
            ws.cell(row=row_idx, column=5, value=asset.remarks or "").border = self.thin_border
            ws.cell(row=row_idx, column=6, value="√" if asset.authenticity else "").border = self.thin_border
            ws.cell(row=row_idx, column=6).alignment = self.center_align
            ws.cell(row=row_idx, column=7, value="√" if asset.integrity else "").border = self.thin_border
            ws.cell(row=row_idx, column=7).alignment = self.center_align
            ws.cell(row=row_idx, column=8, value="√" if asset.non_repudiation else "").border = self.thin_border
            ws.cell(row=row_idx, column=8).alignment = self.center_align
            ws.cell(row=row_idx, column=9, value="√" if asset.confidentiality else "").border = self.thin_border
            ws.cell(row=row_idx, column=9).alignment = self.center_align
            ws.cell(row=row_idx, column=10, value="√" if asset.availability else "").border = self.thin_border
            ws.cell(row=row_idx, column=10).alignment = self.center_align
            ws.cell(row=row_idx, column=11, value="√" if asset.authorization else "").border = self.thin_border
            ws.cell(row=row_idx, column=11).alignment = self.center_align

        # Set column widths
        widths = [12, 25, 15, 15, 30, 10, 10, 12, 10, 10, 10]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

    def _create_attack_tree_sheet(self, project: Project):
        """Create the attack tree analysis sheet."""
        ws = self.wb.create_sheet("攻击树分析", 3)

        ws.cell(row=1, column=1, value="攻击树分析").font = Font(bold=True, size=14)
        ws.cell(row=3, column=1, value="[攻击树图将在此处插入]")

    def _create_tara_result_sheet(self, threats: List[ThreatScenario], assets: List[Asset]):
        """Create the TARA analysis results sheet."""
        ws = self.wb.create_sheet("TARA分析结果", 4)

        # Create asset lookup
        asset_map = {a.id: a for a in assets}

        # Headers - organized by groups
        header_groups = [
            ("资产识别", ["资产ID", "资产名称", "细分类", "分类"]),
            ("威胁&损害场景", ["安全属性", "STRIDE模型", "潜在威胁和损害场景"]),
            ("威胁分析", ["攻击路径", "来源", "WP29映射", "攻击向量", "攻击复杂度", "权限需求", "用户交互", "可行性"]),
            ("影响分析", ["Safety", "Financial", "Operational", "Privacy", "影响等级"]),
            ("风险评估", ["风险等级"]),
            ("风险处置", ["处置决策"]),
            ("风险缓解", ["安全目标", "安全需求", "WP29控制"]),
        ]

        # Write group headers
        col = 1
        for group_name, columns in header_groups:
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(columns) - 1)
            cell = ws.cell(row=1, column=col, value=group_name)
            cell.font = self.header_font
            cell.fill = self.subheader_fill
            cell.border = self.thin_border
            cell.alignment = self.center_align
            col += len(columns)

        # Write column headers
        col = 1
        for group_name, columns in header_groups:
            for header in columns:
                cell = ws.cell(row=2, column=col, value=header)
                cell.font = self.header_font
                cell.fill = self.header_fill
                cell.border = self.thin_border
                cell.alignment = self.center_align
                col += 1

        # Data rows
        for row_idx, threat in enumerate(threats, 3):
            asset = asset_map.get(threat.asset_id)
            mitigation = threat.mitigations[0] if threat.mitigations else None

            # Asset identification
            col = 1
            ws.cell(row=row_idx, column=col, value=asset.asset_id if asset else "").border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=asset.name if asset else "").border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=asset.subcategory if asset else "").border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=asset.category if asset else "").border = self.thin_border
            col += 1

            # Threat scenario
            ws.cell(row=row_idx, column=col, value=threat.security_attribute).border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=threat.stride_type).border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=threat.threat_description).border = self.thin_border
            ws.cell(row=row_idx, column=col).alignment = self.left_align
            col += 1

            # Threat analysis
            ws.cell(row=row_idx, column=col, value=threat.attack_path or "").border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=threat.source_reference or "").border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=threat.wp29_mapping or "").border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=threat.attack_vector or "").border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=threat.attack_complexity or "").border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=threat.privileges_required or "").border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=threat.user_interaction or "").border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=threat.attack_feasibility or "").border = self.thin_border
            col += 1

            # Impact analysis
            ws.cell(row=row_idx, column=col, value=threat.impact_safety or "").border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=threat.impact_financial or "").border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=threat.impact_operational or "").border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=threat.impact_privacy or "").border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=threat.impact_level or "").border = self.thin_border
            col += 1

            # Risk assessment - apply color
            risk_cell = ws.cell(row=row_idx, column=col, value=threat.risk_level_label or "")
            risk_cell.border = self.thin_border
            risk_cell.alignment = self.center_align
            if threat.risk_level and threat.risk_level in self.risk_colors:
                risk_cell.fill = self.risk_colors[threat.risk_level]
            col += 1

            # Risk treatment
            ws.cell(row=row_idx, column=col, value=threat.treatment_decision or "").border = self.thin_border
            col += 1

            # Risk mitigation
            ws.cell(row=row_idx, column=col, value=mitigation.security_goal if mitigation else "").border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=mitigation.security_requirement if mitigation else "").border = self.thin_border
            col += 1
            ws.cell(row=row_idx, column=col, value=mitigation.wp29_control_mapping if mitigation else "").border = self.thin_border

        # Set column widths
        widths = [
            10, 20, 12, 12,  # Asset
            12, 10, 40,  # Threat scenario
            30, 15, 12, 12, 10, 10, 10, 10,  # Threat analysis
            8, 8, 10, 8, 10,  # Impact
            10,  # Risk
            10,  # Treatment
            30, 30, 15,  # Mitigation
        ]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header rows
        ws.freeze_panes = 'A3'
